import datetime
import argparse
import json
import numpy as np
import os
import cv2
import openpyxl as xl
import xlrd
import sys
from align_custom import AlignCustom
from face_feature import FaceFeature
from mtcnn_detect import MTCNNDetect
from tf_graph import FaceRecGraph
from datetime import date
from xlsxwriter.utility import xl_rowcol_to_cell
from openpyxl import Workbook

parser = argparse.ArgumentParser()
parser.add_argument("--mode", type=str, help="Add New Face data", default="input")
args = parser.parse_args(sys.argv[1:])
FRGraph = FaceRecGraph()
aligner = AlignCustom()
extract_feature = FaceFeature(FRGraph)
face_detect = MTCNNDetect(FRGraph, scale_factor=2)
result = "Unknown"


class Recog:
    def __init__(self, mode, new_name, turma):
        self.mode = mode
        self.new_name = new_name
        self.turma = turma

        if (self.mode == "camera"):
            self.camera_recog(turma)
        elif (self.mode == "input"):
            self.create_manual_data(mode, new_name)
        else:
            raise ValueError("Unimplemented mode")


    def create_manual_data(self, mode, new_name):
        self.mode = mode
        self.new_name = new_name

        vs = cv2.VideoCapture(0)  # get input from webcam
        f = open('./facerec_128D.txt', 'r')
        data_set = json.loads(f.read())
        person_imgs = {"Left": [], "Right": [], "Center": []}
        person_features = {"Left": [], "Right": [], "Center": []}
        print("Please start turning slowly. Press 'q' to save and add this new user to the dataset")
        f.close()


        while True:
            _, frame = vs.read()
            rects, landmarks = face_detect.detect_face(frame, 80)  # min face size is set to 80x80
            for (i, rect) in enumerate(rects):
                aligned_frame, pos = aligner.align(160, frame, landmarks[:, i])
                if len(aligned_frame) == 160 and len(aligned_frame[0]) == 160:
                    person_imgs[pos].append(aligned_frame)
                    cv2.imshow("Captured face", aligned_frame)
            key = cv2.waitKey(1) & 0xFF
            if key == ord("q"):
                cv2.destroyAllWindows()
                vs.release()
                break

        for pos in person_imgs:  # there r some exceptions here, but I'll just leave it as this to keep it simple
            person_features[pos] = [np.mean(extract_feature.get_features(person_imgs[pos]), axis=0).tolist()]
        data_set[new_name] = person_features
        f = open('./facerec_128D.txt', 'w')
        f.write(json.dumps(data_set))
        f.close()



    def camera_recog(self, turma):
        self.turma = turma

        filename = 'Presenca.xlsx'

        try:
            wb = xl.load_workbook(filename)
            ws = wb.active
            now = datetime.datetime.now()
            today = date.today()
            data = today.strftime("%d/%m/%y")
            mes = now.strftime('%m')
            dia = now.strftime('%d')
            hora = now.strftime("%H")
            hora_int = int(hora)
            ws.cell(column=1, row=1, value='Nome')
            ws.cell(column=2, row=1, value=data)
            # ws.cell(column=3, row=1, value='Turma')
            wb.save(filename)

        except FileNotFoundError:
            wb = Workbook()
            wb.save('D:/FaceRec-master/Presenca.xlsx')
            ws = wb.active

            now = datetime.datetime.now()
            today = date.today()
            data = today.strftime("%d/%m/%y")
            mes = now.strftime('%m')
            dia = now.strftime('%d')
            hora = now.strftime("%H")
            hora_int = int(hora)
            ws.cell(column=1, row=1, value='Nome')
            ws.cell(column=2, row=1, value=data)
            # ws.cell(column=3, row=1, value='Turma')
            wb.save(filename)


        if hora_int >0 and hora_int <12:
            periodo = 'manha'

        elif hora_int >=12 and hora_int <=18:
            periodo = 'tarde'

        else:
            periodo = 'noite'



        print("[INFO] camera sensor warming up...")
        vs = cv2.VideoCapture(0) #get input from webcam
        frames = 1
        persona = ""

        while True:

            _,frame = vs.read()


            #u can certainly add a roi here but for the sake of a demo i'll just leave it as simple as this
            rects, landmarks = face_detect.detect_face(frame,80)#min face size is set to 80x80
            aligns = []
            positions = []

            for (i, rect) in enumerate(rects):
                aligned_face, face_pos = aligner.align(160,frame,landmarks[:,i])
                if len(aligned_face) == 160 and len(aligned_face[0]) == 160:
                    aligns.append(aligned_face)
                    positions.append(face_pos)
                else:
                    print("Align face failed") #log
            if(len(aligns) > 0):
                features_arr = extract_feature.get_features(aligns)
                recog_data = Recog.findPeople(features_arr, positions)
                for (i, rect) in enumerate(rects):
                    cv2.rectangle(frame,(rect[0],rect[1]),(rect[2],rect[3]),(255,0,0)) #draw bounding box for the face
                    cv2.putText(frame,recog_data[i][0]+" - "+str(recog_data[i][1])+"%",(rect[0],rect[1]),
                                cv2.FONT_HERSHEY_SIMPLEX,1,(255,255,255),1,cv2.LINE_AA)


            cv2.imshow("Frame",frame)
            key = cv2.waitKey(1) & 0xFF

            myPath = r'D:/FaceRec-master/'+filename
            if result != "Unknown" and frames <= 5:
                for sh in xlrd.open_workbook(myPath).sheets():
                    found = Recog.findSpecificPerson(sh, result)

                    if found != -1:
                        frames = 1
                        continue
                    else:
                        frames += 1

                    rv, frame = vs.read()
                    if rv and found == -1 and persona != result and frames == 5: #if the person is NOT in the sheet
                        # (found=-1) and the person is different from last iteraction and you have 5 frames of the same
                        # person, write the name and save the img.
                        #frames = 1
                        ws.cell(column=1, row=ws.max_row+1, value=result)
                        ws.cell(column=2, row=ws.max_row, value='Presente')
                        #ws.cell(column=3, row=ws.max_row, value=turma)
                        wb.save(filename)

                        path = r'D:/FaceRec-master/imgs'

                        cv2.imwrite(os.path.join(path,'%s'%result+'_'+'%s'%dia+'_'+'%s'%mes+'_'+'%s'%periodo+'.jpg'),
                                    frame)
                        print(result, ' foi salvo em ', filename)
                        frames = 1
                        persona = result
                        continue


            if key == ord("q"):
                cv2.destroyAllWindows() #clear all windows
                vs.release() #release the camera
                wb.close()
                break


    def findSpecificPerson(sh, result):
        for row in range(sh.nrows):
            for col in range(sh.ncols):
                myCell = sh.cell(row, col)
                if myCell.value == result:
                    return xl_rowcol_to_cell(row, col)
        return -1

    def findPeople(features_arr, positions, thres = 0.6, percent_thres = 65):
        '''
        :param features_arr: a list of 128d Features of all faces on screen
        :param positions: a list of face position types of all faces on screen
        :param thres: distance threshold
        :return: person name and percentage
        '''

        f = open('./facerec_128D.txt','r')
        data_set = json.loads(f.read())
        returnRes = []


        for (i,features_128D) in enumerate(features_arr):
            global result


            smallest = sys.maxsize
            for person in data_set.keys():
                person_data = data_set[person][positions[i]]
                for data in person_data:
                    distance = np.sqrt(np.sum(np.square(data-features_128D)))
                    if(distance < smallest):
                        smallest = distance
                        result = person

            percentage =  min(100, 100 * thres / smallest)

            if percentage <= percent_thres :
                result = "Unknown"


            returnRes.append((result,percentage))

        return returnRes

class Clear: #clear the person and its coordenates from the txt
    def __init__(self, RemoverNome):
        self.RemoverNome = RemoverNome

        f = open('./facerec_128D.txt', 'r') #read
        g = open('./facerec_128D_new.txt', 'w') #write

        with g as dest_file:
            with f as source_file:
                for line in source_file:
                    element = json.loads(line.strip())
                    if RemoverNome in element:
                        del element[RemoverNome]
                    dest_file.write(json.dumps(element))
        f.close()
        g.close()

        os.remove('./facerec_128D.txt')
        os.rename('./facerec_128D_new.txt','./facerec_128D.txt')

        # for row in ws['A2:B70']: #clean presence spreadsheet
        #     for cell in row:
        #         cell.value = None
        #
        # print('Apagado!')
        # wb.save(filename)
        # cv2.destroyAllWindows()
        # vs.release()
        # break






