import copy
import datetime
import json
import os
import sys
import openpyxl as xl
import imghdr
import ntpath
import imutils

import cv2
import numpy as np

from face_feature import FaceFeature
from tf_graph import FaceRecGraph
from align_custom import AlignCustom
from mtcnn_detect import MTCNNDetect
from datetime import date
from openpyxl import Workbook
from tkinter import messagebox

FRGraph = FaceRecGraph()
extract_feature = FaceFeature(FRGraph)
aligner = AlignCustom()
face_detect = MTCNNDetect(FRGraph, scale_factor=2)


class RecognizePerson:

    def create_manual_data(self, image, new_name):
        f = open(r'./facerec_ICA.txt', 'r')
        data_set = json.loads(f.read())
        person_imgs = {"Left": [], "Right": [], "Center": []}
        person_features = {"Left": [], "Right": [], "Center": []}
        f.close()

        try:
            imghdr.what(new_name)  # check if it's an image
            string = ntpath.basename(new_name)
            new_name = string[:string.index('.')]

            # loop over the rotation angles ensuring no part of the image is cut off
            images_add = []
            for angle in np.arange(-15, 15, 1):
                rotated = imutils.rotate_bound(image, angle)
                cv2.imshow("Rotated", rotated)
                images_add.append(rotated)
                # cv2.destroyAllWindows()
                key = cv2.waitKey(0) & 0xFF
                if key == ord('q'):
                    cv2.destroyAllWindows()

            images_add_copy = copy.deepcopy(images_add)
            images_complete = images_add_copy
            images_flip = []


            for i in images_add_copy:
                image_array = cv2.flip(i, 1)  # flip horizontally

                images_flip.append(image_array)
                # cv2.imshow('Pic', image_array)
                # key = cv2.waitKey(0) & 0xFF
                # if key == ord('q'):
                #     cv2.destroyAllWindows()

            images_complete.extend(images_flip)  # add rotated and flipped images to images_complete

            for images_add_single in images_complete:
                rects, landmarks = face_detect.detect_face(images_add_single, 80)  # min face size is set to 80x80
                for (i, rect) in enumerate(rects):
                    aligned_frame, pos = aligner.align(160, images_add_single, landmarks[:, i])
                    if len(aligned_frame) == 160 and len(aligned_frame[0]) == 160:
                        person_imgs[pos].append(aligned_frame)

            for pos in person_imgs:  # there r some exceptions here, but I'll just leave it as this to keep it simple
                person_features[pos] = [
                    np.mean(extract_feature.get_features(person_imgs[pos]), axis=0).tolist()]
            data_set[new_name] = person_features
            f = open(r'./facerec_ICA.txt', 'w')
            f.write(json.dumps(data_set))
            f.close()
            messagebox.showinfo("Adicionar Usuário", "Usuário adicionado!")

        except IOError:
            # if it's not an image
            vs = cv2.VideoCapture(0)  # get input from webcam
            print(
                "Por favor, comece a mexer a cabeça devagar. Aperte 'q' para salvar e "
                "adicionar o novo usuário ao dataset.")

            while True:
                _, frame = vs.read()
                rects, landmarks = face_detect.detect_face(frame, 80)  # min face size is set to 80x80
                for (i, rect) in enumerate(rects):
                    aligned_frame, pos = aligner.align(160, frame, landmarks[:, i])
                    if len(aligned_frame) == 160 and len(aligned_frame[0]) == 160:
                        person_imgs[pos].append(aligned_frame)
                        cv2.imshow("Adicionar usuário", aligned_frame)
                key = cv2.waitKey(1) & 0xFF
                if key == ord("q"):
                    cv2.destroyAllWindows()
                    vs.release()

                    break

            for pos in person_imgs:  # there r some exceptions here, but I'll just leave it as this to keep it simple
                person_features[pos] = [
                    np.mean(extract_feature.get_features(person_imgs[pos]), axis=0).tolist()]
            data_set[new_name] = person_features
            f = open(r'./facerec_ICA.txt', 'w')
            f.write(json.dumps(data_set))
            f.close()
            messagebox.showinfo("Adicionar Usuário", "Usuário adicionado!")

    @staticmethod
    def camera_recog(frame, frames, result, ret):
        filename = 'Entrada.xlsx'
        path = r'D:/FaceRec-master/imgs'

        frames_to_recog = copy.deepcopy(frame)  # create a deep copy of the frame to avoid modifications
        # from anti spoofing

        try:
            wb = xl.load_workbook(filename)

            now = datetime.datetime.now()
            today = date.today()
            data = today.strftime("%d/%m/%y")
            mes = now.strftime('%m')
            dia = now.strftime('%d')
            hora = now.strftime('%H')
            minutos = now.strftime("%M")

            if dia + '_' + mes in wb.sheetnames:
                ws = wb.get_sheet_by_name(dia + '_' + mes)
                ws.cell(column=1, row=1, value='Nome')
                ws.cell(column=2, row=1, value=data)
                ws.cell(column=3, row=1, value='Hora')
                wb.save(filename)
            else:
                wb.create_sheet(dia + '_' + mes, 0)
                ws = wb.active
                ws.cell(column=1, row=1, value='Nome')
                ws.cell(column=2, row=1, value=data)
                ws.cell(column=3, row=1, value='Hora')
                wb.save(filename)


        except FileNotFoundError:
            wb = Workbook()
            wb.save(r'D:/FaceRec-master/Entrada.xlsx')

            now = datetime.datetime.now()
            today = date.today()
            data = today.strftime("%d/%m/%y")
            mes = now.strftime('%m')
            dia = now.strftime('%d')
            hora = now.strftime('%H')
            minutos = now.strftime("%M")
            wb.create_sheet(dia + '_' + mes, 0)
            std = wb.get_sheet_by_name('Sheet')
            wb.remove_sheet(std)
            ws = wb.active
            ws.cell(column=1, row=1, value='Nome')
            ws.cell(column=2, row=1, value=data)
            ws.cell(column=3, row=1, value='Hora')
            wb.save(filename)

        while True:

            rects, landmarks = face_detect.detect_face(frame, 80)  # min face size is set to 80x80
            aligns = []
            positions = []

            for (i, rect) in enumerate(rects):
                aligned_face, face_pos = aligner.align(160, frame, landmarks[:, i])
                if len(aligned_face) == 160 and len(aligned_face[0]) == 160:
                    aligns.append(aligned_face)
                    positions.append(face_pos)
                else:
                    print("Align face failed")

            if len(aligns) > 0:
                features_arr = extract_feature.get_features(aligns)
                recog_data = RecognizePerson.findPeople(features_arr, positions)
                for (i, rect) in enumerate(rects):
                    cv2.rectangle(frame, (rect[0] - 20, rect[1]), (rect[0] + 150, rect[1] + 150), (255, 0, 0), 2)
                    # bounding box with fixed height and width

                    cv2.putText(frame, recog_data[i][0] + " - " + str(recog_data[i][1]) + "%", (rect[0],
                                                                                                rect[1]),
                                cv2.FONT_HERSHEY_SIMPLEX, 1, (255, 255, 255), 1, cv2.LINE_AA)
                    result = recog_data[-1][0]  # get the last person to recognize

            if result != "Unknown":
                frames += 1

                if ret and frames == 20:  # if it's the same person for 20 frames, then write it on spreadsheet and
                    # save image

                    ws.cell(column=1, row=ws.max_row + 1, value=result)
                    ws.cell(column=2, row=ws.max_row, value='Presente')
                    ws.cell(column=3, row=ws.max_row, value=hora + ':' + minutos)
                    wb.save(filename)

                    cv2.imwrite(os.path.join(path,
                                             '%s' % result + '_' + '%s' % dia + '_' + '%s' % mes + '_' + '%s' % hora
                                             + 'h' + '%s' % minutos + '.jpg'), frames_to_recog)
                    print(result, ' foi salvo em ', filename)

                    text = "True"

                    return text, frames_to_recog, frames, result
                else:
                    text = "False"
                    return text, frames_to_recog, frames, result
            else:
                text = "False"
                return text, frames_to_recog, frames, result

    def findPeople(features_arr, positions, thres=0.6, percent_thres=75):
        """
        :param features_arr: a list of 128d Features of all faces on screen
        :param positions: a list of face position types of all faces on screen
        :param thres: distance threshold
        :return: person name and percentage
        """

        f = open(r'./facerec_ICA.txt', 'r')
        data_set = json.loads(f.read())
        returnRes = []

        for (i, features_ICA) in enumerate(features_arr):
            result = "Unknown"
            smallest = sys.maxsize
            for person in data_set.keys():
                person_data = data_set[person][positions[i]]
                for data in person_data:
                    distance = np.sqrt(np.sum(np.square(data - features_ICA)))
                    if distance < smallest:
                        smallest = distance
                        result = person

            percentage = min(100, 100 * thres / smallest)

            if percentage <= percent_thres:
                result = "Unknown"

            returnRes.append((result, percentage))
        f.close()
        return returnRes


class Clear:  # clear the person coordinates from the txt
    def __init__(self, RemoverNome):
        self.RemoverNome = RemoverNome

        f = open(r'./facerec_ICA.txt', 'r')  # read
        g = open(r'./facerec_ICA_new.txt', 'w')  # write

        with g as dest_file:
            with f as source_file:
                for line in source_file:
                    element = json.loads(line.strip())
                    if RemoverNome in element:
                        del element[RemoverNome]
                    dest_file.write(json.dumps(element))
        f.close()
        g.close()

        os.remove(r'./facerec_ICA.txt')
        os.rename(r'./facerec_ICA_new.txt', r'./facerec_ICA.txt')
