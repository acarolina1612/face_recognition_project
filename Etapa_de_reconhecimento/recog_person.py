import os
import sys
import json
import numpy as np
import cv2
import openpyxl as xl
import xlrd
import threading
from Validacao_de_Pessoas import user_file, welcome_voice
from datetime import datetime, date, timedelta
from xlsxwriter.utility import xl_rowcol_to_cell
from openpyxl import Workbook


class Recog:

    def camera_recog(self, aligner, extract_feature, face_detect):

        global workplace, result, recog_data
        presence_list = 'Presenca.xlsx'

        presence_path = os.path.abspath(os.pardir)  # get the parent folder path. In this case is
        # face_recognition_project
        presence_path = presence_path + '\\face_recognition_project\\Validacao_de_Pessoas\\' + presence_list

        try:
            wb = xl.load_workbook(presence_path)

            now = datetime.now()
            today = date.today()
            today_date = today.strftime("%d/%m/%y")
            month = now.strftime('%m')
            day = now.strftime('%d')

            if day + '_' + month in wb.sheetnames:
                ws = wb.get_sheet_by_name(day + '_' + month)
                ws.cell(column=1, row=1, value='Nome')
                ws.cell(column=2, row=1, value=today_date)
                ws.cell(column=3, row=1, value='Hora')
                ws.cell(column=4, row=1, value='Empresa')
                wb.save(presence_path)
            else:
                wb.create_sheet(day + '_' + month, 0)
                ws = wb.active
                ws.cell(column=1, row=1, value='Nome')
                ws.cell(column=2, row=1, value=today_date)
                ws.cell(column=3, row=1, value='Hora')
                ws.cell(column=4, row=1, value='Empresa')
                wb.save(presence_path)

        except FileNotFoundError:
            wb = Workbook()
            wb.save(presence_path)

            now = datetime.now()
            today = date.today()
            today_date = today.strftime("%d/%m/%y")
            month = now.strftime('%m')
            day = now.strftime('%d')
            wb.create_sheet(day + '_' + month, 0)
            std = wb.get_sheet_by_name('Sheet')
            wb.remove_sheet(std)
            ws = wb.active
            ws.cell(column=1, row=1, value='Nome')
            ws.cell(column=2, row=1, value=today_date)
            ws.cell(column=3, row=1, value='Hora')
            ws.cell(column=4, row=1, value='Empresa')
            wb.save(presence_path)

        vs = cv2.VideoCapture(0)  # get input from webcam
        frames = 1
        time_recognized = datetime.now()
        persona = ''
        person_and_time = [['', datetime.now()], ['', datetime.now()]]
        count_people_found = 0

        while True:

            now = datetime.now()
            month = now.strftime('%m')
            day = now.strftime('%d')
            hour = now.strftime("%H")
            minutes = now.strftime("%M")

            _, frame = vs.read()

            rects, landmarks = face_detect.detect_face(frame, 40)  # min face size is set to 80x80
            aligns = []
            positions = []

            for (i, rect) in enumerate(rects):
                aligned_face, face_pos = aligner.align(160, frame, landmarks[:, i])
                if len(aligned_face) == 160 and len(aligned_face[0]) == 160:
                    aligns.append(aligned_face)
                    positions.append(face_pos)
                else:
                    print("Align face failed")  # log
            if len(aligns) > 0:
                features_arr = extract_feature.get_features(aligns)
                recog_data, company = Recog.findPeople(features_arr,
                                                       positions)  # receive people's coordinates and class
                workplace = company[0]

                for (i, rect) in enumerate(rects):
                    cv2.rectangle(frame, (rect[0], rect[1]), (rect[2], rect[3]), (255, 0, 0), 2)  # draw bounding
                    # box for the face
                    cv2.putText(frame, recog_data[i][0] + " - " + str(recog_data[i][1]) + "%", (rect[0], rect[1]),
                                cv2.FONT_HERSHEY_SIMPLEX, 1, (255, 255, 255), 1, cv2.LINE_AA)

            cv2.imshow("Frame", frame)
            key = cv2.waitKey(1) & 0xFF

            myPath = presence_path

            if result != 'Desconhecido' and frames <= 5 and \
                    (((time_recognized + timedelta(seconds=20)) < datetime.now()) or (person_and_time[0][0] != result
                     and person_and_time[1][0] != result)):
                for sh in xlrd.open_workbook(myPath).sheets():
                    if sh.name == day + '_' + month:  # when the search is in today's sheet, find the person
                        times_found, cell_found, column, row = Recog.findSpecificPerson(sh, result)

                        frames += 1

                        rv, frame = vs.read()

                        if rv and frames == 5:

                            if times_found < 2:
                                ws.cell(column=1, row=ws.max_row + 1, value=result)
                                ws.cell(column=2, row=ws.max_row, value='Presente')
                                ws.cell(column=3, row=ws.max_row, value=hour + ':' + minutes)
                                ws.cell(column=4, row=ws.max_row, value=workplace)
                                wb.save(presence_path)

                            else:  # found more than 2 times, person is leaving
                                ws.cell(column=column + 1, row=row + 1, value=result)
                                ws.cell(column=column + 2, row=row + 1, value='Presente')
                                ws.cell(column=column + 3, row=row + 1, value=hour + ':' + minutes)
                                ws.cell(column=column + 4, row=row + 1, value=workplace)
                                wb.save(presence_path)

                            path = os.path.abspath(os.pardir)  # get the parent folder path. In this case is
                            # face_recognition_project
                            path = path + '\\face_recognition_project\\Validacao_de_Pessoas\\imgs'

                            cv2.imwrite(os.path.join(path,
                                                     '%s' % result + '_' + '%s' % day + '_' + '%s' % month + '_' + '%s' % hour
                                                     + 'h' + '%s' % minutes + '.jpg'), frame)
                            print(result, ' foi salvo em ', presence_list)
                            frames = 1

                            welcome_msg = '\n' + 'Bem vindo, ' + result + '!'
                            info = 'Empresa: ' + workplace

                            t2 = threading.Thread(target=welcome_voice.thread_voice, args=(welcome_msg,))
                            t2.start()

                            user_file.thread_file(welcome_msg, info, result)

                            time_recognized = datetime.now()
                            persona = result
                            count_people_found += 1

                            if count_people_found == 1 and person_and_time[0][0] != persona:
                                person_and_time[0][0] = persona
                                person_and_time[0][1] = time_recognized

                            if count_people_found == 2 and person_and_time[1][0] != persona:
                                person_and_time[1][0] = persona
                                person_and_time[1][1] = time_recognized
                                count_people_found = 0

                            result = 'Desconhecido'

                            continue
                    else:
                        continue

            if key == ord("q"):
                cv2.destroyAllWindows()  # clear all windows
                vs.release()  # release the camera
                wb.close()  # close workbook
                break

    def findSpecificPerson(sh, result):
        times_found = 0
        cell_found = 0
        column = 0
        row = 0
        for row in range(sh.nrows):
            for col in range(sh.ncols):
                myCell = sh.cell(row, col)
                if myCell.value == result:
                    times_found += 1
                    cell_found = xl_rowcol_to_cell(row, col)
                    column = col
                    row = row

        return times_found, cell_found, column, row

    def findPeople(features_arr, positions, thres=0.6, percent_thres=78):
        '''
        :param features_arr: a list of 128d Features of all faces on screen
        :param positions: a list of face position types of all faces on screen
        :param thres: distance threshold
        :return: person name and percentage
        '''
        filename = 'coordinates.txt'

        coordinates_path = os.path.abspath(os.pardir)  # get the parent folder path. In this case is
        # face_recognition_project
        coordinates_path = coordinates_path + '\\face_recognition_project\\Etapa_de_reconhecimento\\' + filename

        f = open(coordinates_path, 'r')
        data_set = json.loads(f.read())
        returnRes = []

        for (i, features_128D) in enumerate(features_arr):
            global result, company

            smallest = sys.maxsize
            for person in data_set.keys():
                person_data = data_set[person][positions[i]]
                for data in person_data:
                    distance = np.sqrt(np.sum(np.square(data - features_128D)))
                    if distance < smallest:
                        smallest = distance
                        result = person
                        company = data_set[person]["Class"]

            percentage = min(100, 100 * thres / smallest)

            if percentage <= percent_thres:
                result = "Desconhecido"
            round_percent = round(percentage)  # round recognize percentage to nearest integer
            returnRes.append((result, round_percent))

        return returnRes, company
