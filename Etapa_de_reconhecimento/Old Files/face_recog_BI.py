import os
import sys
import tkinter as tk
import json
import datetime
import argparse
import numpy as np
import cv2
import openpyxl as xl
import xlrd
import imghdr
import imutils
import ntpath
import copy
import time
import re
import pyttsx3
import threading
from tkinter import ttk, messagebox, font
from align_custom import AlignCustom
from face_feature import FaceFeature
from mtcnn_detect import MTCNNDetect
from tf_graph import FaceRecGraph
from datetime import date
from xlsxwriter.utility import xl_rowcol_to_cell
from openpyxl import Workbook
from tkinter import filedialog
from PIL import Image, ImageTk
from playsound import playsound

NORM_FONT = ("Verdana", 10)
LARGE_FONT = ("Verdana", 12)

parser = argparse.ArgumentParser()
parser.add_argument("--mode", type=str, help="Add New Face data", default="input")
args = parser.parse_args(sys.argv[1:])
FRGraph = FaceRecGraph()
aligner = AlignCustom()
extract_feature = FaceFeature(FRGraph)
face_detect = MTCNNDetect(FRGraph, scale_factor=2)
result = "Unknown"


class App(tk.Tk):

    def __init__(self, *args, **kwargs):  # arguments and keyword arguments

        tk.Tk.__init__(self, *args, **kwargs)
        tk.Tk.iconbitmap(self, default="faceicon.ico")
        tk.Tk.wm_title(self, "Reconhecimento facial")

        self.geometry("800x400")
        self.resizable(0, 0)
        container = tk.Frame(self)

        container.pack(side="top", fill="both", expand=True)

        self.frames = {}

        for F in (StartPage, PageOne):
            frame = F(container, self)

            self.frames[F] = frame

            frame.grid(row=0, column=2, sticky="nsew")

        self.show_frame(StartPage)

    def show_frame(self, cont):
        frame = self.frames[cont]
        frame.tkraise()


class StartPage(tk.Frame):

    def __init__(self, parent, controller):

        tk.Frame.__init__(self, parent)

        insercaoERemocao = tk.Label(self, text='Cadastro/Remoção de usuários', width=0, height=0, padx=10, pady=12,
                                    font=LARGE_FONT)
        insercaoERemocao.grid(row=0, column=0, stick='nsew')
        insercaoERemocaoFrame = ttk.Frame(self, width=400, height=360, relief='raised')
        insercaoERemocaoFrame.grid(row=1, column=0, sticky="nsew")

        button = ttk.Button(insercaoERemocaoFrame, text="Adicionar usuário",
                            command=lambda: controller.show_frame(PageOne))
        button.pack(padx=150, pady=100)

        def popupmsg(msg):
            popup = tk.Tk()
            popup.wm_title("Remover Usuário")

            label = ttk.Label(popup, text=msg, font=NORM_FONT)
            label.pack(side="top", fill="x", pady=10)

            frm = tk.Frame(popup, borderwidth=1)

            f = open(r'./facerec_128D.txt', 'r')
            data_set = json.loads(f.read())
            Nomes = []

            for key, value in data_set.items():
                Nomes.append(key)
            f.close()

            Nome = ttk.Combobox(popup, values=sorted(Nomes))
            Nome.pack(pady=15, padx=20)
            Nome.focus_set()

            def next_step():
                if Nome.get():
                    # the user entered data in the mandatory entry: proceed to next step
                    f = open(r'./facerec_128D.txt', 'r')
                    data_set = json.loads(f.read())
                    Nomes = []

                    for key, value in data_set.items():
                        Nomes.append(key)
                    f.close()

                    RemoverNome = Nome.get()
                    popup.destroy()

                    Clear(RemoverNome)
                    messagebox.showinfo("Remover Usuário", "Usuário removido!")

                else:
                    # the mandatory field is empty
                    messagebox.showinfo("Remover Usuário", "Selecione um usuário!")
                    Nome.focus_set()
                    popup.destroy()
                    popupmsg(msg)

            B1 = ttk.Button(frm, text="OK", command=lambda: next_step())
            B1.pack(side='left')

            B2 = ttk.Button(frm, text="Cancelar", command=popup.destroy)
            B2.pack(side='left')

            frm.pack(side='bottom')

            popup.mainloop()

        button2 = ttk.Button(insercaoERemocaoFrame, text="Remover usuário",
                             command=lambda: popupmsg("Selecione o nome do usuário que será removido"))
        button2.pack(padx=0, pady=10)

        ReconhecimentoUsuarios = tk.Label(self, text='Reconhecimento de usuários', width=0, height=0, padx=10, pady=12,
                                          font=LARGE_FONT)
        ReconhecimentoUsuarios.grid(row=0, column=1, stick='nsew')
        ReconhecimentoUsuariosFrame = ttk.Frame(self, width=400, height=360, relief='raised')
        ReconhecimentoUsuariosFrame.grid(row=1, column=1, sticky="nsew")

        button3 = ttk.Button(ReconhecimentoUsuariosFrame, text="Reconhecer usuário",
                             command=lambda: Recog("camera", "", "", ""))
        button3.pack(padx=150, pady=165)


class PageOne(tk.Frame):

    def __init__(self, parent, controller):
        ttk.Frame.__init__(self, parent)

        bframe = tk.Frame(self)
        cframe = tk.Frame(self)

        Nome = tk.Label(bframe, text="Nome e sobrenome", width=0, height=0, padx=40, pady=50,
                        font=LARGE_FONT)
        Nome.grid(row=0, column=0, stick='nsew')
        NomeFrame = ttk.Frame(bframe, width=600, height=360)
        NomeFrame.grid(row=1, column=0)

        NomeFunc = tk.Entry(NomeFrame)
        NomeFunc.pack(pady=10, padx=10)
        NomeFunc.focus_set()

        Turma = tk.Label(bframe, text="Empresa", width=0, height=0, padx=10, pady=30,
                         font=LARGE_FONT)
        Turma.grid(row=3, column=0, stick='nsew')
        TurmaFrame = ttk.Frame(bframe, width=400, height=360)
        TurmaFrame.grid(row=4, column=0)

        Combo = ttk.Combobox(TurmaFrame, values=["ICA", "PUC-Rio", "Externo"])
        Combo.set("ICA")
        Combo.pack(pady=10, padx=10)
        Combo.focus_set()

        def add_pic(new_name, turma):
            path = filedialog.askopenfile(title='Selecione a foto',
                                          filetypes=[('Image Files', ['.jpeg', '.jpg', '.JPG', '.png', '.gif',
                                                                      '.tiff', '.tif', '.bmp'])])

            if not path:
                PageOne("", "")

            else:
                im = cv2.imread(path.name)
                # cv2.imshow('Foto', im)
                Recog.create_manual_data("", "", path.name, im, turma)
                controller.show_frame(StartPage)

        def add_manual(new_name, turma):
            Recog.create_manual_data("", "input", new_name, "", turma)
            controller.show_frame(StartPage)

        def next_step():
            if NomeFunc.get():
                # the user entered data in the mandatory entry: proceed to next step
                new_name = NomeFunc.get()
                NomeFunc.delete(0, 'end')
                turma = Combo.get()

                # add person in dataset
                popup2 = tk.Tk()
                popup2.wm_title("Adicionar Usuário")

                label = ttk.Label(popup2, text='Selecione uma foto ou pegue as coordenadas do usuário',
                                  font=NORM_FONT)
                label.pack(side="top", fill="x", pady=10)

                frm = tk.Frame(popup2, borderwidth=1)

                B1 = ttk.Button(frm, text="Inserir foto",
                                command=lambda: [add_pic(new_name, turma), popup2.destroy()])
                B1.pack(side='left')

                B2 = ttk.Button(frm, text="Pegar coordenadas",
                                command=lambda: [add_manual(new_name, turma), popup2.destroy()])
                B2.pack(side='right')

                frm.pack(side='bottom', pady='15')

                popup2.mainloop()
                controller.show_frame(StartPage)

            else:
                # the mandatory field is empty
                messagebox.showinfo("Adicionar Usuário", "Escreva o nome do usuário!")
                NomeFunc.focus_set()

        Ok = ttk.Button(TurmaFrame, text='OK', command=lambda: next_step())
        Ok.pack(sid='left')

        Cancelar = ttk.Button(TurmaFrame, text='Cancelar',
                              command=lambda: [controller.show_frame(StartPage), NomeFunc.delete(0, 'end')])
        Cancelar.pack(side='left')

        bframe.pack(side='left')

        AdicionarUsuario = tk.Label(cframe, text='Adicionar Usuário', width=0, height=0, padx=0, pady=0,
                                    font=LARGE_FONT)
        AdicionarUsuario.grid(row=0, column=3)
        AdicionarUsuarioFrame = ttk.Frame(cframe, width=400, height=300, relief='sunken')
        AdicionarUsuarioFrame.grid(row=1, column=3, padx=50)

        cframe.pack(side='right')


class Recog:
    def __init__(self, mode, new_name, image, turma):
        self.mode = mode
        self.new_name = new_name
        self.turma = turma

        if (self.mode == "camera"):
            self.camera_recog()
        elif (self.mode == "input"):
            self.create_manual_data(mode, new_name, image, turma)
        else:
            raise ValueError("Unimplemented mode")

    def create_manual_data(self, mode, new_name, image, turma):

        f = open(r'./facerec_128D.txt', 'r')
        data_set = json.loads(f.read())
        person_imgs = {"Left": [], "Right": [], "Center": []}
        person_features = {"Left": [], "Right": [], "Center": [], "Class": []}
        f.close()

        try:
            imghdr.what(new_name)  # check if it's an image
            string = ntpath.basename(new_name)
            new_name = string[:string.index('.')]  # get the name before "."

            # loop over the rotation angles ensuring no part of the image is cut off
            images_add = []

            if re.match("^[a-zA-Z0-9 _]*$", new_name):  # check if the image name does not have special characters
                for angle in np.arange(-15, 15, 1):
                    rotated = imutils.rotate_bound(image, angle)
                    # cv2.imshow("Rotated", rotated)
                    images_add.append(rotated)
                    # cv2.destroyAllWindows()
                    # key = cv2.waitKey(0) & 0xFF
                    # if key == ord('q'):
                    #     cv2.destroyAllWindows()

                images_add_copy = copy.deepcopy(images_add)
                images_complete = images_add_copy
                images_flip = []

                for i in images_add_copy:
                    image_array = cv2.flip(i, 1)  # flip horizontally

                    images_flip.append(image_array)
                    # cv2.imshow('Pic', image_array)
                    # key = cv2.waitKey(0) & 0xFF
                    # if key == ord('q'):
                    #     cv2.destroyAllWindows()

                images_complete.extend(images_flip)  # add rotated and flipped images to images_complete

                for images_add_single in images_complete:
                    rects, landmarks = face_detect.detect_face(images_add_single, 80)  # min face size is set to 80x80
                    for (i, rect) in enumerate(rects):
                        aligned_frame, pos = aligner.align(160, images_add_single, landmarks[:, i])
                        if len(aligned_frame) == 160 and len(aligned_frame[0]) == 160:
                            person_imgs[pos].append(aligned_frame)

                for pos in person_imgs:  # there r some exceptions here, but I'll just leave it as this to keep it simple
                    person_features[pos] = [
                        np.mean(extract_feature.get_features(person_imgs[pos]), axis=0).tolist()]
                data_set[new_name] = person_features
                data_set[new_name]["Class"] = [turma]
                f = open(r'./facerec_128D.txt', 'w')
                f.write(json.dumps(data_set))
                f.close()
                cv2.destroyAllWindows()
                messagebox.showinfo("Adicionar Usuário", "Usuário adicionado!")

            else:  # if the image name has special characters
                messagebox.showwarning('Aviso', 'Não insira caracteres especiais no nome da foto!')

        except IOError:  # get the coordinates manually

            vs = cv2.VideoCapture(1)  # get input from webcam
            print(
                "Por favor, comece a mexer a cabeça devagar. Aperte 'q' para salvar e "
                "adicionar o novo usuário ao dataset.")

            while True:
                _, frame = vs.read()
                rects, landmarks = face_detect.detect_face(frame, 80)  # min face size is set to 80x80
                for (i, rect) in enumerate(rects):
                    aligned_frame, pos = aligner.align(160, frame, landmarks[:, i])
                    if len(aligned_frame) == 160 and len(aligned_frame[0]) == 160:
                        person_imgs[pos].append(aligned_frame)
                        cv2.imshow("Adicionar usuário", aligned_frame)
                key = cv2.waitKey(1) & 0xFF
                if key == ord("q"):
                    cv2.destroyAllWindows()
                    vs.release()

                    break

            for pos in person_imgs:  # there r some exceptions here, but I'll just leave it as this to keep it simple
                person_features[pos] = [
                    np.mean(extract_feature.get_features(person_imgs[pos]), axis=0).tolist()]
            data_set[new_name] = person_features
            data_set[new_name]["Class"] = [turma]
            f = open(r'./facerec_128D.txt', 'w')
            f.write(json.dumps(data_set))
            f.close()
            messagebox.showinfo("Adicionar Usuário", "Usuário adicionado!")

    def camera_recog(self):

        global empresa
        filename = 'Presenca.xlsx'

        try:
            wb = xl.load_workbook(filename)

            now = datetime.datetime.now()
            today = date.today()
            data = today.strftime("%d/%m/%y")
            mes = now.strftime('%m')
            dia = now.strftime('%d')

            if dia + '_' + mes in wb.sheetnames:
                ws = wb.get_sheet_by_name(dia + '_' + mes)
                ws.cell(column=1, row=1, value='Nome')
                ws.cell(column=2, row=1, value=data)
                ws.cell(column=3, row=1, value='Hora')
                ws.cell(column=4, row=1, value='Empresa')
                wb.save(filename)
            else:
                wb.create_sheet(dia + '_' + mes, 0)
                ws = wb.active
                ws.cell(column=1, row=1, value='Nome')
                ws.cell(column=2, row=1, value=data)
                ws.cell(column=3, row=1, value='Hora')
                ws.cell(column=4, row=1, value='Empresa')
                wb.save(filename)

        except FileNotFoundError:
            wb = Workbook()
            wb.save(r'C:/FaceRec-master/Presenca.xlsx')

            now = datetime.datetime.now()
            today = date.today()
            data = today.strftime("%d/%m/%y")
            mes = now.strftime('%m')
            dia = now.strftime('%d')
            wb.create_sheet(dia + '_' + mes, 0)
            std = wb.get_sheet_by_name('Sheet')
            wb.remove_sheet(std)
            ws = wb.active
            ws.cell(column=1, row=1, value='Nome')
            ws.cell(column=2, row=1, value=data)
            ws.cell(column=3, row=1, value='Hora')
            ws.cell(column=4, row=1, value='Empresa')
            wb.save(filename)

        vs = cv2.VideoCapture(1)  # get input from webcam
        frames = 1
        persona = ""

        while True:

            now = datetime.datetime.now()
            mes = now.strftime('%m')
            dia = now.strftime('%d')
            hora = now.strftime("%H")
            minutos = now.strftime("%M")

            _, frame = vs.read()

            rects, landmarks = face_detect.detect_face(frame, 80)  # min face size is set to 80x80
            aligns = []
            positions = []

            for (i, rect) in enumerate(rects):
                aligned_face, face_pos = aligner.align(160, frame, landmarks[:, i])
                if len(aligned_face) == 160 and len(aligned_face[0]) == 160:
                    aligns.append(aligned_face)
                    positions.append(face_pos)
                else:
                    print("Align face failed")  # log
            if len(aligns) > 0:
                features_arr = extract_feature.get_features(aligns)
                recog_data, turma = Recog.findPeople(features_arr, positions)  # receive people's coordinates and class
                empresa = turma[0]

                for (i, rect) in enumerate(rects):
                    cv2.rectangle(frame, (rect[0], rect[1]), (rect[2], rect[3]), (255, 0, 0), 2)  # draw bounding
                    # box for the face
                    cv2.putText(frame, recog_data[i][0] + " - " + str(recog_data[i][1]) + "%", (rect[0], rect[1]),
                                cv2.FONT_HERSHEY_SIMPLEX, 1, (255, 255, 255), 1, cv2.LINE_AA)

            cv2.imshow("Frame", frame)
            key = cv2.waitKey(1) & 0xFF

            myPath = r'C:/FaceRec-master/' + filename
            if result != "Unknown" and frames <= 5:
                for sh in xlrd.open_workbook(myPath).sheets():
                    if sh.name == dia + '_' + mes:  # when the search is in today's sheet, find the person
                        found = Recog.findSpecificPerson(sh, result)

                        if found != -1:  # if the person is not found in today's spreadsheet,
                            # put frames back to 1
                            frames = 1
                            continue
                        else:
                            frames += 1

                        rv, frame = vs.read()
                        if rv and found == -1 and persona != result and frames == 5:  # if the person is NOT in the sheet
                            # (found=-1) and the person is different from last interaction and you have 5 frames of the same
                            # person, write the name and save the img.

                            ws.cell(column=1, row=ws.max_row + 1, value=result)
                            ws.cell(column=2, row=ws.max_row, value='Presente')
                            ws.cell(column=3, row=ws.max_row, value=hora + ':' + minutos)
                            ws.cell(column=4, row=ws.max_row, value=empresa)
                            wb.save(filename)

                            path = r'C:/FaceRec-master/imgs'

                            cv2.imwrite(os.path.join(path,
                                                     '%s' % result + '_' + '%s' % dia + '_' + '%s' % mes + '_' + '%s' % hora
                                                     + 'h' + '%s' % minutos + '.jpg'), frame)
                            print(result, ' foi salvo em ', filename)
                            frames = 1
                            persona = result

                            saudacao = '\n' + 'Bem vindo, ' + result + '!'
                            info = 'Empresa: ' + empresa

                            def thread_voice(saudacao):  # say 'welcome' to the user
                                playsound(r'./audio3.mp3')
                                engine = pyttsx3.init()
                                volume = engine.getProperty('volume')
                                engine.setProperty('volume', volume + 0.5)
                                engine.setProperty('rate', 200)
                                engine.say(saudacao)
                                engine.runAndWait()

                            def thread_file(saudacao):  # create customer's file
                                w = tk.Toplevel()
                                w.iconbitmap(default="faceicon.ico")
                                w.wm_title('Ficha do usuário')
                                w.geometry('600x600')
                                w.resizable(0, 0)
                                logo = ImageTk.PhotoImage(file=path + '/' + '%s' % result + '_' + '%s' % dia + '_' +
                                                               '%s' % mes + '_' + '%s' % hora
                                                               + 'h' + '%s' % minutos + '.jpg')

                                label_right = tk.Label(w)

                                w1 = tk.Label(label_right,
                                              padx=70,
                                              text=info,
                                              font=('Helvetica', 16)).pack(side='bottom')
                                w2 = tk.Label(label_right, image=logo, width=450, height=450, relief='sunken').pack(
                                    side="bottom")
                                w3 = tk.Label(label_right,
                                              text=saudacao,
                                              font=('Helvetica', 20)).pack(side='top')
                                label_right.pack(anchor='center')
                                w.update()
                                time.sleep(3)
                                w.destroy()

                            t1 = threading.Thread(target=thread_voice, args=(saudacao,))
                            t1.start()

                            thread_file(saudacao)

                            continue
                    else:
                        continue

            if key == ord("q"):
                cv2.destroyAllWindows()  # clear all windows
                vs.release()  # release the camera
                wb.close()
                break

    def findSpecificPerson(sh, result):
        for row in range(sh.nrows):
            for col in range(sh.ncols):
                myCell = sh.cell(row, col)
                if myCell.value == result:
                    return xl_rowcol_to_cell(row, col)
        return -1

    def findPeople(features_arr, positions, thres=0.6, percent_thres=78):
        '''
        :param features_arr: a list of 128d Features of all faces on screen
        :param positions: a list of face position types of all faces on screen
        :param thres: distance threshold
        :return: person name and percentage
        '''

        f = open(r'./facerec_128D.txt', 'r')
        data_set = json.loads(f.read())
        returnRes = []

        for (i, features_128D) in enumerate(features_arr):
            global result, turma

            smallest = sys.maxsize
            for person in data_set.keys():
                person_data = data_set[person][positions[i]]
                for data in person_data:
                    distance = np.sqrt(np.sum(np.square(data - features_128D)))
                    if distance < smallest:
                        smallest = distance
                        result = person
                        turma = data_set[person]["Class"]

            percentage = min(100, 100 * thres / smallest)

            if percentage <= percent_thres:
                result = "Unknown"

            returnRes.append((result, percentage))

        return returnRes, turma


class Clear:  # clear the person and its coordinates from the txt
    def __init__(self, RemoverNome):
        self.RemoverNome = RemoverNome

        f = open(r'./facerec_128D.txt', 'r')  # read
        g = open(r'./facerec_128D_new.txt', 'w')  # write

        with g as dest_file:
            with f as source_file:
                for line in source_file:
                    element = json.loads(line.strip())
                    if RemoverNome in element:
                        del element[RemoverNome]
                    dest_file.write(json.dumps(element))
        f.close()
        g.close()

        os.remove(r'./facerec_128D.txt')
        os.rename(r'./facerec_128D_new.txt', r'./facerec_128D.txt')


app = App()
app.mainloop()
app.bind('<Escape>', sys.exit())
