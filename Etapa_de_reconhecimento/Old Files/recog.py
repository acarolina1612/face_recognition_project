import datetime
import time
import argparse
import sys
import json
import numpy as np
import os
import cv2
import openpyxl
import xlsxwriter
import xlrd
from align_custom import AlignCustom
from face_feature import FaceFeature
from mtcnn_detect import MTCNNDetect
from tf_graph import FaceRecGraph
from datetime import date
from xlsxwriter.utility import xl_rowcol_to_cell
from openpyxl import Workbook

result = "Unknown"

def main(args):
    mode = args.mode
    if(mode == "camera"):
        camera_recog()
    elif (mode == "input"):
        create_manual_data()
    else:
        raise ValueError("Unimplemented mode")

def camera_recog():


    filename = 'Presenca.xlsx'

    try:
        wb = openpyxl.load_workbook(filename)
        ws = wb.active
        now = datetime.datetime.now()
        today = date.today()
        data = today.strftime("%d/%m/%y")
        mes = now.strftime('%m')
        dia = now.strftime('%d')
        hora = now.strftime("%H")
        hora_int = int(hora)
        ws.cell(column=1, row=1, value='Nome')
        ws.cell(column=2, row=1, value=data)
        wb.save(filename)

    except FileNotFoundError:
        wb = Workbook()
        wb.save('C:/Users/GabrielTaranto/FaceRec-master/Presenca.xlsx')
        ws = wb.active

        now = datetime.datetime.now()
        today = date.today()
        data = today.strftime("%d/%m/%y")
        mes = now.strftime('%m')
        dia = now.strftime('%d')
        hora = now.strftime("%H")
        hora_int = int(hora)
        ws.cell(column=1, row=1, value='Nome')
        ws.cell(column=2, row=1, value=data)
        wb.save(filename)


    if hora_int >0 and hora_int <12:
        periodo = 'manha'

    elif hora_int >=12 and hora_int <=18:
        periodo = 'tarde'

    else:
        periodo = 'noite'



    print("[INFO] camera sensor warming up...")
    vs = cv2.VideoCapture(0) #get input from webcam
    detect_time = time.time()
    frames = 1

    while True:

        _,frame = vs.read()


        #u can certainly add a roi here but for the sake of a demo i'll just leave it as simple as this
        rects, landmarks = face_detect.detect_face(frame,80)#min face size is set to 80x80
        aligns = []
        positions = []

        for (i, rect) in enumerate(rects):
            aligned_face, face_pos = aligner.align(160,frame,landmarks[:,i])
            if len(aligned_face) == 160 and len(aligned_face[0]) == 160:
                aligns.append(aligned_face)
                positions.append(face_pos)
            else:
                print("Align face failed") #log
        if(len(aligns) > 0):
            features_arr = extract_feature.get_features(aligns)
            recog_data = findPeople(features_arr, positions)
            for (i, rect) in enumerate(rects):
                cv2.rectangle(frame,(rect[0],rect[1]),(rect[2],rect[3]),(255,0,0)) #draw bounding box for the face
                cv2.putText(frame,recog_data[i][0]+" - "+str(recog_data[i][1])+"%",(rect[0],rect[1]),cv2.FONT_HERSHEY_SIMPLEX,1,(255,255,255),1,cv2.LINE_AA)
                frames += 1

        cv2.imshow("Frame",frame)
        key = cv2.waitKey(1) & 0xFF


        myPath = r'C:/Users/GabrielTaranto/FaceRec-master/'+filename
        for sh in xlrd.open_workbook(myPath).sheets():
            if result != "Unknown":
                found = findCell(sh, result)
                rv, frame = vs.read()

                if rv and (found == "" or found == -1) and frames == 10: #if found is empty or -1 and you have
                    # 10 frames of the same person, write the name and save the img.
                    ws.cell(column=1, row=ws.max_row+1, value=result)
                    ws.cell(column=2, row=ws.max_row, value='Presente')
                    wb.save(filename)

                    path = r'C:/Users/GabrielTaranto/FaceRec-master/imgs'

                    cv2.imwrite(os.path.join(path,'%s'%result+'%s'%dia+'_'+'%s'%mes+'_'+'%s'%periodo+'.jpg'), frame)
                    print(result, ' foi salvo em ', filename)
                    frames = 1
                    continue


        if key == ord("c"):

            for row in ws['A2:B70']:
                for cell in row:
                    cell.value = None

            print('Apagado!')
            wb.save(filename)
            cv2.destroyAllWindows()
            vs.release()
            break


        elif key == ord("q"):
            cv2.destroyAllWindows() #clear all windows
            vs.release() #release the camera

            break


def findCell(sh, result):
    for row in range(sh.nrows):
        for col in range(sh.ncols):
            myCell = sh.cell(row, col)
            if myCell.value == result:
                return xl_rowcol_to_cell(row, col)
    return -1

def findPeople(features_arr, positions, thres = 0.6, percent_thres = 70):
    '''
    :param features_arr: a list of 128d Features of all faces on screen
    :param positions: a list of face position types of all faces on screen
    :param thres: distance threshold
    :return: person name and percentage
    '''

    f = open('./facerec_128D.txt','r')
    data_set = json.loads(f.read())
    returnRes = []


    for (i,features_128D) in enumerate(features_arr):
        global result


        smallest = sys.maxsize
        for person in data_set.keys():
            person_data = data_set[person][positions[i]]
            for data in person_data:
                distance = np.sqrt(np.sum(np.square(data-features_128D)))
                if(distance < smallest):
                    smallest = distance
                    result = person

        percentage =  min(100, 100 * thres / smallest)

        if percentage <= percent_thres :
            result = "Unknown"


        returnRes.append((result,percentage))

    return returnRes


if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument("--mode", type=str, help="Run camera recognition", default="camera")
    args = parser.parse_args(sys.argv[1:])
    FRGraph = FaceRecGraph()
    MTCNNGraph = FaceRecGraph()
    aligner = AlignCustom()
    extract_feature = FaceFeature(FRGraph)
    face_detect = MTCNNDetect(MTCNNGraph, scale_factor=2) #scale_factor, rescales image for faster detection
    main(args)