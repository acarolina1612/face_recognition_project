import os
import sys
import json
import datetime
import numpy as np
import cv2
import openpyxl as xl
import xlrd
import imghdr
import imutils
import ntpath
import copy
import re
import threading
import user_file
import welcome_voice
import Cadastro_de_pessoas.load_models as load_models
from datetime import date
from xlsxwriter.utility import xl_rowcol_to_cell
from openpyxl import Workbook
from Interface import Interface, loading_bar
from tkinter import messagebox

t_1 = threading.Thread(target=loading_bar.main)
t_1.start()

parser, args, FRGraph, aligner, extract_feature, face_detect = load_models.main()

Interface.main()


class Recog:
    def __init__(self, mode, new_name, image, company):
        self.mode = mode
        self.new_name = new_name
        self.company = company

        if (self.mode == "camera"):
            self.camera_recog()
        elif (self.mode == "input"):
            self.create_manual_data(mode, new_name, image, company)
        else:
            raise ValueError("Unimplemented mode")

    def create_manual_data(self, mode, new_name, image, company):

        f = open(r'./coordinates.txt', 'r')
        data_set = json.loads(f.read())
        person_imgs = {"Left": [], "Right": [], "Center": []}
        person_features = {"Left": [], "Right": [], "Center": [], "Class": []}
        f.close()

        try:
            imghdr.what(new_name)  # check if it's an image
            string = ntpath.basename(new_name)
            new_name = string[:string.index('.')]  # get the name before "."

            # loop over the rotation angles ensuring no part of the image is cut off
            images_add = []

            if re.match("^[a-zA-Z0-9 _]*$", new_name):  # check if the image name does not have special characters
                for angle in np.arange(-15, 15, 1):
                    rotated = imutils.rotate_bound(image, angle)
                    # cv2.imshow("Rotated", rotated)
                    images_add.append(rotated)
                    # cv2.destroyAllWindows()
                    # key = cv2.waitKey(0) & 0xFF
                    # if key == ord('q'):
                    #     cv2.destroyAllWindows()

                images_add_copy = copy.deepcopy(images_add)
                images_complete = images_add_copy
                images_flip = []

                for i in images_add_copy:
                    image_array = cv2.flip(i, 1)  # flip horizontally

                    images_flip.append(image_array)
                    # cv2.imshow('Pic', image_array)
                    # key = cv2.waitKey(0) & 0xFF
                    # if key == ord('q'):
                    #     cv2.destroyAllWindows()

                images_complete.extend(images_flip)  # add rotated and flipped images to images_complete

                for images_add_single in images_complete:
                    rects, landmarks = face_detect.detect_face(images_add_single, 80)  # min face size is set to 80x80
                    for (i, rect) in enumerate(rects):
                        aligned_frame, pos = aligner.align(160, images_add_single, landmarks[:, i])
                        if len(aligned_frame) == 160 and len(aligned_frame[0]) == 160:
                            person_imgs[pos].append(aligned_frame)

                for pos in person_imgs:  # there r some exceptions here, but I'll just leave it as this to keep it simple
                    person_features[pos] = [
                        np.mean(extract_feature.get_features(person_imgs[pos]), axis=0).tolist()]
                data_set[new_name] = person_features
                data_set[new_name]["Class"] = [company]
                f = open(r'./coordinates.txt', 'w')
                f.write(json.dumps(data_set))
                f.close()
                cv2.destroyAllWindows()
                messagebox.showinfo("Adicionar Usuário", "Usuário adicionado!")

            else:  # if the image name has special characters
                messagebox.showwarning('Aviso', 'Não insira caracteres especiais no nome da foto!')

        except IOError:  # get the coordinates manually

            vs = cv2.VideoCapture(0)  # get input from webcam
            print(
                "Por favor, comece a mexer a cabeça devagar. Aperte 'q' para salvar e "
                "adicionar o novo usuário ao dataset.")

            while True:
                _, frame = vs.read()
                rects, landmarks = face_detect.detect_face(frame, 80)  # min face size is set to 80x80
                for (i, rect) in enumerate(rects):
                    aligned_frame, pos = aligner.align(160, frame, landmarks[:, i])
                    if len(aligned_frame) == 160 and len(aligned_frame[0]) == 160:
                        person_imgs[pos].append(aligned_frame)
                        cv2.imshow("Adicionar usuário", aligned_frame)
                key = cv2.waitKey(1) & 0xFF
                if key == ord("q"):
                    cv2.destroyAllWindows()
                    vs.release()

                    break

            for pos in person_imgs:  # there r some exceptions here, but I'll just leave it as this to keep it simple
                person_features[pos] = [
                    np.mean(extract_feature.get_features(person_imgs[pos]), axis=0).tolist()]
            data_set[new_name] = person_features
            data_set[new_name]["Class"] = [company]
            f = open(r'./coordinates.txt', 'w')
            f.write(json.dumps(data_set))
            f.close()
            messagebox.showinfo("Adicionar Usuário", "Usuário adicionado!")

    def camera_recog(self):

        global workplace, result, recog_data
        filename = 'Presenca.xlsx'
        presence_path = r'./'+filename

        try:
            wb = xl.load_workbook(presence_path)

            now = datetime.datetime.now()
            today = date.today()
            today_date = today.strftime("%d/%m/%y")
            month = now.strftime('%m')
            day = now.strftime('%d')

            if day + '_' + month in wb.sheetnames:
                ws = wb.get_sheet_by_name(day + '_' + month)
                ws.cell(column=1, row=1, value='Nome')
                ws.cell(column=2, row=1, value=today_date)
                ws.cell(column=3, row=1, value='Hora')
                ws.cell(column=4, row=1, value='Empresa')
                wb.save(presence_path)
            else:
                wb.create_sheet(day + '_' + month, 0)
                ws = wb.active
                ws.cell(column=1, row=1, value='Nome')
                ws.cell(column=2, row=1, value=today_date)
                ws.cell(column=3, row=1, value='Hora')
                ws.cell(column=4, row=1, value='Empresa')
                wb.save(presence_path)

        except FileNotFoundError:
            wb = Workbook()
            wb.save(presence_path)

            now = datetime.datetime.now()
            today = date.today()
            today_date = today.strftime("%d/%m/%y")
            month = now.strftime('%m')
            day = now.strftime('%d')
            wb.create_sheet(day + '_' + month, 0)
            std = wb.get_sheet_by_name('Sheet')
            wb.remove_sheet(std)
            ws = wb.active
            ws.cell(column=1, row=1, value='Nome')
            ws.cell(column=2, row=1, value=today_date)
            ws.cell(column=3, row=1, value='Hora')
            ws.cell(column=4, row=1, value='Empresa')
            wb.save(presence_path)

        vs = cv2.VideoCapture(0)  # get input from webcam
        frames = 1
        # persona = ""

        while True:

            now = datetime.datetime.now()
            month = now.strftime('%m')
            day = now.strftime('%d')
            hour = now.strftime("%H")
            minutes = now.strftime("%M")

            _, frame = vs.read()

            rects, landmarks = face_detect.detect_face(frame, 40)  # min face size is set to 80x80
            aligns = []
            positions = []

            for (i, rect) in enumerate(rects):
                aligned_face, face_pos = aligner.align(160, frame, landmarks[:, i])
                if len(aligned_face) == 160 and len(aligned_face[0]) == 160:
                    aligns.append(aligned_face)
                    positions.append(face_pos)
                else:
                    print("Align face failed")  # log
            if len(aligns) > 0:
                features_arr = extract_feature.get_features(aligns)
                recog_data, company = Recog.findPeople(features_arr, positions)  # receive people's coordinates and class
                workplace = company[0]

                for (i, rect) in enumerate(rects):
                    cv2.rectangle(frame, (rect[0], rect[1]), (rect[2], rect[3]), (255, 0, 0), 2)  # draw bounding
                    # box for the face
                    cv2.putText(frame, recog_data[i][0] + " - " + str(recog_data[i][1]) + "%", (rect[0], rect[1]),
                                cv2.FONT_HERSHEY_SIMPLEX, 1, (255, 255, 255), 1, cv2.LINE_AA)

            cv2.imshow("Frame", frame)
            key = cv2.waitKey(1) & 0xFF

            myPath = presence_path

            if result != 'Desconhecido' and frames <= 5:
                for sh in xlrd.open_workbook(myPath).sheets():
                    if sh.name == day + '_' + month:  # when the search is in today's sheet, find the person
                        times_found, cell_found, column, row = Recog.findSpecificPerson(sh, result)

                        frames += 1

                        rv, frame = vs.read()

                        if rv and frames == 5:

                            if times_found < 2:
                                ws.cell(column=1, row=ws.max_row + 1, value=result)
                                ws.cell(column=2, row=ws.max_row, value='Presente')
                                ws.cell(column=3, row=ws.max_row, value=hour + ':' + minutes)
                                ws.cell(column=4, row=ws.max_row, value=workplace)
                                wb.save(presence_path)

                            else:  # found more than 2 times, person is leaving
                                ws.cell(column=column + 1, row=row + 1, value=result)
                                ws.cell(column=column + 2, row=row + 1, value='Presente')
                                ws.cell(column=column + 3, row=row + 1, value=hour + ':' + minutes)
                                ws.cell(column=column + 4, row=row + 1, value=workplace)
                                wb.save(presence_path)

                            path = r'./imgs'

                            cv2.imwrite(os.path.join(path,
                                                     '%s' % result + '_' + '%s' % day + '_' + '%s' % month + '_' + '%s' % hour
                                                     + 'h' + '%s' % minutes + '.jpg'), frame)
                            print(result, ' foi salvo em ', filename)
                            frames = 1

                            welcome_msg = '\n' + 'Bem vindo, ' + result + '!'
                            info = 'Empresa: ' + workplace

                            t2 = threading.Thread(target=welcome_voice.thread_voice, args=(welcome_msg,))
                            t2.start()

                            user_file.thread_file(welcome_msg, info, result)
                            result = 'Desconhecido'

                            continue
                    else:
                        continue

            if key == ord("q"):
                cv2.destroyAllWindows()  # clear all windows
                vs.release()  # release the camera
                wb.close()  # close workbook
                break

    def findSpecificPerson(sh, result):
        times_found = 0
        cell_found = 0
        column = 0
        row = 0
        for row in range(sh.nrows):
            for col in range(sh.ncols):
                myCell = sh.cell(row, col)
                if myCell.value == result:
                    times_found += 1
                    cell_found = xl_rowcol_to_cell(row, col)
                    column = col
                    row = row

        return times_found, cell_found, column, row

    def findPeople(features_arr, positions, thres=0.6, percent_thres=78):
        '''
        :param features_arr: a list of 128d Features of all faces on screen
        :param positions: a list of face position types of all faces on screen
        :param thres: distance threshold
        :return: person name and percentage
        '''

        f = open(r'./coordinates.txt', 'r')
        data_set = json.loads(f.read())
        returnRes = []

        for (i, features_128D) in enumerate(features_arr):
            global result, company

            smallest = sys.maxsize
            for person in data_set.keys():
                person_data = data_set[person][positions[i]]
                for data in person_data:
                    distance = np.sqrt(np.sum(np.square(data - features_128D)))
                    if distance < smallest:
                        smallest = distance
                        result = person
                        company = data_set[person]["Class"]

            percentage = min(100, 100 * thres / smallest)

            if percentage <= percent_thres:
                result = "Desconhecido"
            round_percent = round(percentage)  # round recognize percentage to nearest integer
            returnRes.append((result, round_percent))

        return returnRes, company
